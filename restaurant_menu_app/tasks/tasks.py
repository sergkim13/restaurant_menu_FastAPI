import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from restaurant_menu_app.db.main_db import crud
from restaurant_menu_app.tasks.tasks_config import celery_app


@celery_app.task
async def create_task():
    return "Hello"


async def get_all_from_database(db: AsyncSession) -> dict:
    result = await crud.get_all(db)
    return result


async def save_to_xlsx(data: dict):
    menus = data
    book = openpyxl.Workbook()
    sheet = book.active
    menu_row, menu_column, menu_counter = 1, 1, 1
    for menu in menus:
        sheet.cell(menu_row, menu_column).value = menu_counter
        sheet.cell(menu_row, menu_column + 1).value = menu["menu_title"]
        sheet.cell(menu_row, menu_column + 2).value = menu["menu_description"]
        menu_counter += 1
        submenu_row = menu_row + 1
        submenu_column = menu_column + 1
        submenu_counter = 1
        for submenu in menu["child_submenus"]:
            sheet.cell(submenu_row, submenu_column).value = submenu_counter
            sheet.cell(submenu_row, submenu_column + 1).value = submenu["submenu_title"]
            sheet.cell(submenu_row, submenu_column + 2).value = submenu["submenu_description"]
            submenu_counter += 1
            dish_row = submenu_row + 1
            dish_column = submenu_column + 1
            dish_counter = 1
            for dish in submenu["child_dishes"]:
                sheet.cell(dish_row, dish_column).value = dish_counter
                sheet.cell(dish_row, dish_column + 1).value = dish["dish_title"]
                sheet.cell(dish_row, dish_column + 2).value = dish["dish_description"]
                sheet.cell(dish_row, dish_column + 3).value = dish["dish_price"]
                dish_counter += 1
                dish_row += 1
            submenu_row = dish_row
        menu_row = submenu_row
    book.save("book.xlsx")
    book.close()
